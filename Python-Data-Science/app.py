import sqlite3
from datetime import datetime, timedelta
import pandas as pd
from flask import Flask, render_template, request, jsonify, send_file, g
import openpyxl
from openpyxl.styles import Font, Alignment
import os
import shutil
import io

app = Flask(__name__)

# مسار قاعدة البيانات النسبي (يعمل على Render وReplit)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(BASE_DIR, "overtime_records.db")

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DB_FILE)
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def init_db():
    with app.app_context():
        conn = get_db()
        c = conn.cursor()
        c.execute("""
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            job_title TEXT,
            salary REAL
        )
        """)
        c.execute("""
        CREATE TABLE IF NOT EXISTS overtime_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER,
            date TEXT,
            start_time TEXT,
            end_time TEXT,
            hours REAL,
            multiplier REAL,
            overtime_amount REAL,
            notes TEXT,
            period_id INTEGER,
            created_at TEXT,
            FOREIGN KEY(employee_id) REFERENCES employees(id)
        )
        """)
        c.execute("""
        CREATE TABLE IF NOT EXISTS periods (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE,
            start_date TEXT,
            end_date TEXT,
            is_open INTEGER,
            year INTEGER DEFAULT 2025
        )
        """)
        default_emps = [("عادل حسام","موظف",220),
                        ("محمد إسماعيل عبدالرحيم","موظف",270),
                        ("عتيق الزمان أنور","موظف",180)]
        for emp in default_emps:
            try:
                c.execute("INSERT INTO employees(name, job_title, salary) VALUES (?,?,?)", emp)
            except sqlite3.IntegrityError:
                pass
        
        c.execute("SELECT * FROM periods WHERE is_open=1")
        if not c.fetchall():
            today = datetime.now().strftime("%Y-%m-%d")
            current_year = datetime.now().year
            c.execute("INSERT INTO periods(name,start_date,end_date,is_open,year) VALUES (?,?,?,?,?)",
                      (f"الفترة المفتوحة الأولى {today}", today, today, 1, current_year))
        conn.commit()

def get_employees():
    c = get_db().cursor()
    c.execute("SELECT id, name, job_title, salary FROM employees ORDER BY name")
    rows = c.fetchall()
    return rows

def add_employee(name, job_title, salary):
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("INSERT INTO employees(name, job_title, salary) VALUES (?,?,?)",
                  (name, job_title, salary))
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False

def update_employee(emp_id, job_title, salary):
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE employees SET job_title=?, salary=? WHERE id=?",
              (job_title, salary, emp_id))
    conn.commit()

def delete_employee(emp_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM employees WHERE id=?", (emp_id,))
    c.execute("DELETE FROM overtime_records WHERE employee_id=?", (emp_id,))
    conn.commit()

def save_record(employee_id, date, start, end, hours, multiplier, amount, notes, period_id):
    conn = get_db()
    c = conn.cursor()
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c.execute("""
    INSERT INTO overtime_records
    (employee_id,date,start_time,end_time,hours,multiplier,overtime_amount,notes,period_id,created_at)
    VALUES (?,?,?,?,?,?,?,?,?,?)
    """,(employee_id,date,start,end,hours,multiplier,amount,notes,period_id,created_at))
    conn.commit()

def update_record(record_id, start, end, hours, multiplier, amount, notes):
    conn = get_db()
    c = conn.cursor()
    c.execute("""
    UPDATE overtime_records
    SET start_time=?, end_time=?, hours=?, multiplier=?, overtime_amount=?, notes=?
    WHERE id=?
    """, (start,end,hours,multiplier,amount,notes,record_id))
    conn.commit()

def delete_record(record_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM overtime_records WHERE id=?", (record_id,))
    conn.commit()

def fetch_records(employee_id=None, period_id=None):
    c = get_db().cursor()
    if employee_id and period_id:
        c.execute("""
        SELECT id,date,start_time,end_time,multiplier,hours,overtime_amount,notes
        FROM overtime_records
        WHERE employee_id=? AND period_id=?
        ORDER BY date ASC,id ASC
        """,(employee_id, period_id))
    else:
         c.execute("SELECT id,date,start_time,end_time,multiplier,hours,overtime_amount,notes FROM overtime_records ORDER BY date ASC,id ASC")
    rows = c.fetchall()
    return rows

def get_open_period():
    c = get_db().cursor()
    c.execute("SELECT id,name FROM periods WHERE is_open=1")
    row = c.fetchone()
    return row

def close_period(period_id):
    conn = get_db()
    c = conn.cursor()
    today = datetime.now().strftime("%Y-%m-%d")
    current_year = datetime.now().year
    c.execute("UPDATE periods SET is_open=0,end_date=? WHERE id=?", (today, period_id))
    conn.commit()
    c.execute("INSERT INTO periods(name,start_date,end_date,is_open,year) VALUES (?,?,?,?,?)",
              (f"الفترة التالية {today}", today, today, 1, current_year))
    conn.commit()

def fetch_periods():
    c = get_db().cursor()
    c.execute("SELECT id,name,year,is_open,start_date,end_date FROM periods ORDER BY year DESC, id DESC")
    rows = c.fetchall()
    return [(r[0],r[1],r[2]) for r in rows]

def add_period(name, year):
    conn = get_db()
    c = conn.cursor()
    try:
        today = datetime.now().strftime("%Y-%m-%d")
        c.execute("INSERT INTO periods(name,start_date,end_date,is_open,year) VALUES (?,?,?,?,?)",
                  (name, today, today, 1, year))
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False

def backup_db_web():
    backup_folder = "backups"
    if not os.path.exists(backup_folder):
        os.makedirs(backup_folder)
    backup_path = os.path.join(backup_folder,f"overtime_records_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
    shutil.copy(DB_FILE, backup_path)
    return f"تم حفظ النسخة الاحتياطية في: {backup_path}"

def compute_hours(start_time,end_time):
    try:
        fmt = "%H:%M"
        t1 = datetime.strptime(start_time, fmt)
        t2 = datetime.strptime(end_time, fmt)
        if t2 < t1:
            t2 += timedelta(days=1)
        delta = t2 - t1
        return round(delta.total_seconds()/3600,3)
    except:
        return 0.0

def compute_hourly_wage(salary):
    return round(salary/(30*8),3)

def compute_overtime_amount(hours,hourly_wage,multiplier):
    return round(hours*hourly_wage*multiplier,3)

def generate_excel_in_memory(employee_name, records):
    df = pd.DataFrame(records, columns=["التاريخ","من","إلى","مضاعف","الساعات","الساعات الإضافية","الملاحظات"])
    df.insert(0, "التسلسل", range(1, len(df)+1))
    total_hours = df["الساعات"].sum()
    total_amount = df["الساعات الإضافية"].sum()
    total_row = {"التسلسل":"-","التاريخ":"-","من":"-","إلى":"-","مضاعف":"-",
                 "الساعات": total_hours,"الساعات الإضافية": total_amount,"الملاحظات":"-"}
    df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
    
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name="الساعات الإضافية", startrow=2)
    
    wb = writer.book
    ws = writer.sheets["الساعات الإضافية"]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    ws.cell(row=1, column=1).value = f"الساعات الإضافية للموظف: {employee_name}"
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    ws.sheet_view.rightToLeft = True

    writer.close()
    output.seek(0)
    return output

@app.route('/')
def index():
    employees = get_employees()
    periods = fetch_periods()
    return render_template('index.html', employees=employees, periods=periods)

@app.route('/get_data', methods=['POST'])
def get_data():
    data = request.json
    emp_id = data.get('employee_id')
    period_id = data.get('period_id')
    
    if not emp_id or not period_id:
        return jsonify({'error': 'Missing data'}), 400
        
    emp_id = int(emp_id)
    period_id = int(period_id)

    salary = 0.0
    hourly_wage = 0.0
    for emp in get_employees():
        if emp[0] == emp_id:
            salary = emp[3]
            hourly_wage = compute_hourly_wage(salary)
            break
            
    records = fetch_records(emp_id, period_id)
    total_hours = sum(r[5] for r in records)
    total_amount = sum(r[6] for r in records)
    
    return jsonify({
        'salary': salary,
        'hourly_wage': hourly_wage,
        'records': records,
        'total_hours': total_hours,
        'total_amount': total_amount
    })

@app.route('/calculate', methods=['POST'])
def calculate():
    data = request.json
    try:
        hours = compute_hours(data['start'], data['end'])
        hourly_wage = float(data.get('hourly_wage', 0))
        multiplier = float(data.get('multiplier', 1))
        amount = compute_overtime_amount(hours, hourly_wage, multiplier)
        return jsonify({'hours': f"{hours:.3f}", 'amount': f"{amount:.3f}"})
    except Exception as e:
        return jsonify({'hours': '0.000', 'amount': '0.000', 'error': str(e)})

@app.route('/add_record', methods=['POST'])
def add_record_route():
    data = request.json
    try:
        save_record(
            data['employee_id'],
            data['date'],
            data['start_time'],
            data['end_time'],
            float(data['hours']),
            float(data['multiplier']),
            float(data['overtime_amount']),
            data['notes'],
            data['period_id']
        )
        return jsonify({'success': True, 'message': 'تم حفظ السجل'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/update_record', methods=['POST'])
def update_record_route():
    data = request.json
    try:
        hours = compute_hours(data['start_time'], data['end_time'])
        hourly_wage = compute_hourly_wage(float(data['salary']))
        multiplier = float(data['multiplier'])
        amount = compute_overtime_amount(hours, hourly_wage, multiplier)

        update_record(
            data['record_id'],
            data['start_time'],
            data['end_time'],
            hours,
            multiplier,
            amount,
            data['notes']
        )
        return jsonify({'success': True, 'message': 'تم تحديث السجل'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/delete_record', methods=['POST'])
def delete_record_route():
    data = request.json
    try:
        delete_record(data['record_id'])
        return jsonify({'success': True, 'message': 'تم حذف السجل'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/add_employee', methods=['POST'])
def add_employee_route():
    data = request.json
    if add_employee(data['name'], data['job_title'], float(data['salary'])):
        return jsonify({'success': True, 'message': 'تم إضافة الموظف'})
    else:
        return jsonify({'success': False, 'message': 'الموظف موجود مسبقاً'})

@app.route('/get_employee', methods=['POST'])
def get_employee_route():
    emp_id = request.json['id']
    for emp in get_employees():
        if emp[0] == int(emp_id):
            return jsonify({'id': emp[0], 'name': emp[1], 'job_title': emp[2], 'salary': emp[3]})
    return jsonify({'error': 'Employee not found'}), 404

@app.route('/update_employee', methods=['POST'])
def update_employee_route():
    data = request.json
    update_employee(data['id'], data['job_title'], float(data['salary']))
    return jsonify({'success': True, 'message': 'تم تحديث الموظف'})

@app.route('/delete_employee', methods=['POST'])
def delete_employee_route():
    data = request.json
    delete_employee(data['id'])
    return jsonify({'success': True, 'message': 'تم حذف الموظف'})

@app.route('/add_period', methods=['POST'])
def add_period_route():
    data = request.json
    if add_period(data['name'], int(data['year'])):
        return jsonify({'success': True, 'message': 'تم إضافة الفترة'})
    else:
        return jsonify({'success': False, 'message': 'الفترة موجودة مسبقاً'})

@app.route('/close_period', methods=['POST'])
def close_period_route():
    data = request.json
    close_period(data['period_id'])
    new_periods = fetch_periods()
    return jsonify({'success': True, 'message': 'تم إغلاق الفترة', 'periods': new_periods})

@app.route('/export', methods=['POST'])
def export_route():
    data = request.json
    emp_id = data.get('employee_id')
    period_id = data.get('period_id')
    emp_name = data.get('employee_name')
    
    records = fetch_records(emp_id, period_id)
    formatted_records = [(r[1],r[2],r[3],r[4],r[5],r[6],r[7]) for r in records]
    
    excel_buffer = generate_excel_in_memory(emp_name, formatted_records)
    
    return send_file(
        excel_buffer,
        as_attachment=True,
        download_name=f"Overtime_{emp_name}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/backup', methods=['POST'])
def backup_route():
    message = backup_db_web()
    return jsonify({'success': True, 'message': message})

if __name__ == "__main__":
    import os
    init_db()
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=True, host='0.0.0.0', port=port)
